import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from concurrent.futures import ProcessPoolExecutor
from datetime import datetime, time, date, timedelta
from functools import lru_cache
import logging
from utils import interpretar_previsoes, dias_da_semana

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class CartaoPrograma:
    def __init__(self, previsores, crime_data):
        self.previsores = previsores
        self.crime_data = crime_data
        self.pontos_patrulhamento = []
        self._validar_dados()

    def _validar_dados(self):
        required_columns = ["DIA_SEMANA", "HORARIO_FATO", "BAIRRO", "LOGRADOURO", "LATITUDE", "LONGITUDE"]

        if not all(col in self.crime_data.df.columns for col in required_columns):
            missing_cols = [col for col in required_columns if col not in self.crime_data.df.columns]
            raise ValueError(f"Colunas obrigatórias ausentes: {missing_cols}")

        if self.crime_data.df.empty:
            raise ValueError("DataFrame está vazio")

    def gerar_pontos_patrulhamento(self):
        logger.info("Iniciando geração de pontos de patrulhamento")
        self.pontos_patrulhamento = []

        try:
            with ProcessPoolExecutor() as executor:
                futures = [executor.submit(self._processar_dia, dia) for dia in range(7)]
                for future in futures:
                    self.pontos_patrulhamento.extend(future.result())

            self.pontos_patrulhamento.sort(key=lambda x: (x["DIA_SEMANA"], x["HORARIO_INICIO"]))
            logger.info(f"Gerados {len(self.pontos_patrulhamento)} pontos de patrulhamento")
            return self.pontos_patrulhamento

        except Exception as e:
            logger.error(f"Erro na geração de pontos: {e}")
            raise

    def _processar_dia(self, dia_semana):
        pontos_dia = []
        for turno in range(4):
            pontos_dia.extend(self._processar_turno(dia_semana, turno))
        return pontos_dia

    def _processar_turno(self, dia_semana, turno):
        pontos_turno = []
        horario_inicio_turno = turno * 6

        for i in range(6):
            try:
                hora = horario_inicio_turno + i
                bairro = self.crime_data.df.sample(1)["BAIRRO"].iloc[0]

                # Obter a previsão do modelo
                tipo_crime, probabilidade_previsao = self.previsores.prever_local_horario(bairro, dia_semana, hora)
                probabilidade_combinada = self._combinar_probabilidades(
                    self._calcular_probabilidades(bairro, dia_semana, hora), probabilidade_previsao
                )
                objetivo = interpretar_previsoes(tipo_crime, probabilidade_combinada, dia_semana, hora)

                # Calcular o horário de início e término baseado na hora do objetivo
                horario_inicio = datetime.combine(date.today(), time(hora))
                horario_fim = horario_inicio + timedelta(minutes=20)

                ponto = {
                    "DIA_SEMANA": dia_semana,
                    "HORARIO_INICIO": horario_inicio,
                    "HORARIO_TERMINO": horario_fim,
                    "BAIRRO": bairro,
                    "LOGRADOURO": self.crime_data.df[self.crime_data.df["BAIRRO"] == bairro].sample(1)["LOGRADOURO"].iloc[0],
                    "LATITUDE": self.crime_data.df[self.crime_data.df["BAIRRO"] == bairro].sample(1)["LATITUDE"].iloc[0],
                    "LONGITUDE": self.crime_data.df[self.crime_data.df["BAIRRO"] == bairro].sample(1)["LONGITUDE"].iloc[0],
                    "OBJETIVO": objetivo,
                    "MISSAO": f"Patrulhamento preventivo em {bairro}",
                    "OBSERVACAO": ""
                }
                pontos_turno.append(ponto)

            except Exception as e:
                logger.warning(f"Erro ao adicionar ponto {i} para o turno {turno}: {e}")

        return pontos_turno

    @lru_cache(maxsize=128)
    def _calcular_probabilidades(self, bairro, dia_semana, hora):
        try:
            tipo_crime, probabilidade = self.previsores.prever_local_horario(bairro, dia_semana, hora)
            return probabilidade
        except Exception as e:
            logger.error(f"Erro ao calcular probabilidades: {e}")
            raise

    def _combinar_probabilidades(self, probabilidade_historica, probabilidade_previsao):
        return (probabilidade_historica + probabilidade_previsao) / 2

    def gerar_excel(self, filename):
        logger.info(f"Iniciando geração do arquivo Excel: {filename}")

        try:
            wb = openpyxl.Workbook()

            for dia_semana in range(7):
                pontos_dia = [p for p in self.pontos_patrulhamento if p["DIA_SEMANA"] == dia_semana]
                if not pontos_dia:
                    continue

                self._criar_aba_excel(wb, dia_semana, pontos_dia)

            wb.remove(wb["Sheet"])
            wb.save(filename)
            logger.info(f"Arquivo Excel gerado com sucesso: {filename}")
            return filename

        except Exception as e:
            logger.error(f"Erro na geração do arquivo Excel: {e}")
            raise

    def _criar_aba_excel(self, wb, dia_semana, pontos_dia):
        df_patrulhamento = pd.DataFrame(pontos_dia)
        ws = wb.create_sheet(title=dias_da_semana[dia_semana])

        df_patrulhamento["ORDEM_OCUPACAO"] = range(1, len(df_patrulhamento) + 1)
        cols = df_patrulhamento.columns.tolist()
        cols = cols[-1:] + cols[:-1]
        df_patrulhamento = df_patrulhamento[cols]

        for col in ["HORARIO_INICIO", "HORARIO_TERMINO"]:
            df_patrulhamento[col] = pd.to_datetime(df_patrulhamento[col]).dt.strftime('%H:%M')

        for row in range(len(df_patrulhamento.index) + 1):
            for col in range(len(df_patrulhamento.columns)):
                if row == 0:
                    ws.cell(row=row + 1, column=col + 1).value = df_patrulhamento.columns[col]
                else:
                    ws.cell(row=row + 1, column=col + 1).value = df_patrulhamento.iloc[row - 1, col]

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_value = str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        return ws
